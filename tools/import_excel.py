from __future__ import annotations

import datetime as dt
import json
import re
import shutil
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SOURCE = Path(
    r"C:/Users/clopez/OneDrive - TTS Flooring & Countertops/Downloads/copy of Equipment Status.xlsx"
)
LOCAL_COPY = ROOT / "data" / "source-equipment-status.xlsx"
OUTPUT = ROOT / "data" / "equipment-data.json"


def classify(note: object) -> str:
    if note is None or str(note).strip() == "":
        return "unknown"
    text = str(note).lower()
    if "n/a" in text:
        return "inactive"
    if any(word in text for word in ["urgent", "down", "shut down", "not working", "fault"]):
        return "down"
    if any(
        word in text
        for word in ["leak", "issue", "repair", "replaced", "fixed", "service", "maintenance", "pm"]
    ):
        return "maintenance"
    if any(word in text for word in ["operational", "working", "good", "ready"]):
        return "operational"
    return "unknown"


def clean_machine(name: object) -> str:
    return str(name).strip().replace("Compresor", "Compressor")


def machine_id(name: object) -> str:
    return re.sub(r"[^a-z0-9]+", "-", clean_machine(name).lower()).strip("-")


def iso_date(value: object) -> str:
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def category_for(name: str) -> str:
    if "Crane" in name:
        return "Lifting"
    if "Compressor" in name or "Air Dryer" in name or "Water Recycling" in name:
        return "Utilities"
    return "Production"


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    if not LOCAL_COPY.exists():
        shutil.copy2(SOURCE, LOCAL_COPY)

    workbook = load_workbook(LOCAL_COPY, data_only=True, read_only=True)
    status_sheet = workbook["Main WS"]
    header_row = next(status_sheet.iter_rows(min_row=2, max_row=2, max_col=14, values_only=True))
    machine_names = [clean_machine(header) for header in header_row[1:] if header]

    updates: list[dict[str, object]] = []
    for row in status_sheet.iter_rows(min_row=3, max_col=14, values_only=True):
        raw_date = row[0]
        if not raw_date:
            continue
        date = iso_date(raw_date)
        for note, name in zip(row[1:], machine_names):
            if note is None or str(note).strip() == "":
                continue
            updates.append(
                {
                    "id": f"{machine_id(name)}-{date}",
                    "machineId": machine_id(name),
                    "machine": name,
                    "date": date,
                    "status": classify(note),
                    "note": str(note).strip(),
                    "source": "Equipment Status.xlsx",
                }
            )

    locations = [
        "Shop Floor - Bay 1",
        "Shop Floor - Bay 2",
        "Waterjet Area",
        "Warehouse",
        "Compressor Room",
    ]
    machines: list[dict[str, object]] = []
    for index, name in enumerate(machine_names):
        mid = machine_id(name)
        machine_updates = [update for update in updates if update["machineId"] == mid]
        latest = machine_updates[-1] if machine_updates else None
        latest_signal = next(
            (update for update in reversed(machine_updates) if update["status"] != "unknown"),
            latest,
        )
        counts = Counter(str(update["status"]) for update in machine_updates)
        machines.append(
            {
                "id": mid,
                "name": name,
                "category": category_for(name),
                "location": locations[index % len(locations)],
                "model": name,
                "serialNumber": f"TTS-{index + 1:03d}",
                "currentStatus": latest_signal["status"] if latest_signal else "unknown",
                "lastUpdated": latest_signal["date"] if latest_signal else None,
                "latestNote": latest_signal["note"] if latest_signal else "",
                "runtimeTodayHours": round(2.2 + (index * 0.7) % 5.8, 1),
                "utilization": int(42 + (index * 7) % 43),
                "oee": int(56 + (index * 6) % 29),
                "statusCounts": dict(counts),
            }
        )

    demo_overrides = {
        "saber-jet-2-crane": ("down", "Hoist not responding. Create urgent repair work order."),
        "voyager-2": ("maintenance", "PM due soon: 10,000 hour service."),
        "shop-compressor": ("maintenance", "Quarterly service due in 5 days."),
    }
    for machine in machines:
        if machine["id"] in demo_overrides:
            machine["currentStatus"], machine["latestNote"] = demo_overrides[str(machine["id"])]

    forklift_sheet = workbook["Forklifts and Warehouse"]
    forklifts = []
    forklift_values = {}
    for row_number, row in enumerate(
        forklift_sheet.iter_rows(min_row=1, max_row=14, max_col=2, values_only=True),
        start=1,
    ):
        forklift_values[row_number] = list(row)
    for start in [1, 6, 11]:
        unit = forklift_values[start][1]
        if unit:
            forklifts.append(
                {
                    "unit": str(unit),
                    "make": str(forklift_values[start + 1][1] or ""),
                    "model": str(forklift_values[start + 2][1] or ""),
                    "serialNumber": str(forklift_values[start + 3][1] or ""),
                }
            )
    forklifts.append(
        {
            "unit": "4",
            "make": "Hyundai",
            "model": "25LC-7A",
            "serialNumber": "HHKHHC08F0004178",
        }
    )

    payload = {
        "sourceWorkbook": str(SOURCE),
        "generatedAt": dt.datetime.now().isoformat(timespec="seconds"),
        "machines": machines,
        "updates": updates[-2200:],
        "workOrders": [],
        "pmSchedule": [],
        "forklifts": forklifts,
        "technicians": ["Tom R.", "Mike D.", "Sara K.", "Alex P.", "John D."],
        "issueTypes": [
            "Operational check",
            "Leak repaired",
            "Preventive Maintenance",
            "Urgent Repair",
            "Electrical",
            "Hydraulic",
            "Calibration",
            "Parts Replacement",
        ],
    }
    OUTPUT.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT}")
    print(f"{len(machines)} machines, {len(updates)} status updates")


if __name__ == "__main__":
    main()
